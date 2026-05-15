#!/usr/bin/env python3
"""
Patch the 'website' field in existing .md event files using hyperlinks
extracted from the XLSX Tickets column.

Usage:
    python patch_website_field.py --xlsx <path> --dir <events_md_dir>
"""

import argparse
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


# ── Helpers ───────────────────────────────────────────────────────────────────

def slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^\w\s-]", "", text).strip().lower()
    text = re.sub(r"[\s_-]+", "_", text)
    return text[:80]


def build_url_map(xlsx_path: Path) -> dict[str, str]:
    """
    Read the XLSX and return {slug: ticket_url} for every event row
    that has a hyperlink in the Tickets column (col index 9).
    """
    wb = load_workbook(str(xlsx_path), read_only=False)
    ws = wb.active
    url_map = {}

    for row in ws.iter_rows():
        name_cell   = row[0] if len(row) > 0 else None
        ticket_cell = row[9] if len(row) > 9 else None

        name = name_cell.value if name_cell else None
        if not name or not str(name).strip():
            continue

        if ticket_cell and ticket_cell.hyperlink:
            url  = ticket_cell.hyperlink.target
            slug = slugify(str(name).strip())
            url_map[slug] = url

    return url_map


def patch_md_file(md_path: Path, url: str) -> bool:
    """
    Replace the website field value in a .md frontmatter file.
    Returns True if the file was changed, False if already correct or not found.
    """
    text = md_path.read_text(encoding="utf-8")

    # Match the website line inside frontmatter
    pattern = r"(^website: ')(.*?)(')", 
    new_line = f"website: '{url}'"

    # Replace website: '...' with the new URL, only inside the frontmatter block
    fm_match = re.match(r"^(---\s*\n)(.*?)(\n---)", text, re.DOTALL)
    if not fm_match:
        return False

    fm_body = fm_match.group(2)
    new_fm_body = re.sub(
        r"^website: '.*?'",
        f"website: '{url}'",
        fm_body,
        flags=re.MULTILINE,
    )

    if new_fm_body == fm_body:
        return False  # no change

    new_text = text.replace(fm_body, new_fm_body, 1)
    md_path.write_text(new_text, encoding="utf-8")
    return True


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Patch website field in .md files from XLSX ticket hyperlinks."
    )
    parser.add_argument("--xlsx", required=True, help="Path to the source .xlsx file")
    parser.add_argument("--dir",  default="./events_md", help="Directory of .md files")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    md_dir    = Path(args.dir)

    if not xlsx_path.exists():
        print(f"ERROR: XLSX not found: {xlsx_path}")
        return
    if not md_dir.is_dir():
        print(f"ERROR: Directory not found: {md_dir}")
        return

    print(f"Reading hyperlinks from: {xlsx_path}")
    url_map = build_url_map(xlsx_path)
    print(f"Found {len(url_map)} ticket URLs\n")

    md_files = sorted(md_dir.glob("*.md"))
    patched  = 0
    no_url   = []
    unchanged = 0

    for md_path in md_files:
        slug = md_path.stem  # filename without .md
        url  = url_map.get(slug)

        if not url:
            no_url.append(md_path.name)
            print(f"  ⚠️   {md_path.name}  (no ticket URL found)")
            continue

        changed = patch_md_file(md_path, url)
        if changed:
            patched += 1
            print(f"  ✅  {md_path.name}")
            print(f"       → {url}")
        else:
            unchanged += 1
            print(f"  –   {md_path.name}  (already set)")

    print(f"\n{'─'*60}")
    print(f"Patched:   {patched}")
    print(f"Unchanged: {unchanged}")
    print(f"No URL:    {len(no_url)}")
    if no_url:
        print("\nFiles with no matching ticket URL:")
        for name in no_url:
            print(f"  • {name}")


if __name__ == "__main__":
    main()
